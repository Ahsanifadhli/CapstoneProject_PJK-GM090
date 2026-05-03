import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# ============ SHEET 1: TIMELINE ============
ws = wb.active
ws.title = "Timeline Proyek"

headers = ["No","Minggu","Tanggal Mulai","Tanggal Selesai","Fase","Kategori","Task","Deskripsi","PIC","Status","Prioritas","Milestone","Catatan"]

# Colors
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)

CAT_COLORS = {
    "Project Management": PatternFill("solid", fgColor="DBEAFE"),
    "AI/ML": PatternFill("solid", fgColor="FEF3C7"),
    "Backend": PatternFill("solid", fgColor="D1FAE5"),
    "Frontend": PatternFill("solid", fgColor="EDE9FE"),
    "All Team": PatternFill("solid", fgColor="FCE7F3"),
}

WEEK_COLORS = {
    "Weekly #1": PatternFill("solid", fgColor="F0F9FF"),
    "Weekly #2": PatternFill("solid", fgColor="FFF7ED"),
    "Weekly #3": PatternFill("solid", fgColor="F0FDF4"),
    "Weekly #4": PatternFill("solid", fgColor="FDF4FF"),
    "Weekly #5": PatternFill("solid", fgColor="FFF1F2"),
}

data = [
    [1,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","Project Management","Setup GitHub Repository","Inisialisasi repo private; buat struktur folder; setup .gitignore; buat branch strategy","Project Manager","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [2,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","Project Management","Koordinasi Tim Minggu 1","Memimpin weekly sync perdana; bahas pembagian tugas; setup Google Calendar","Project Manager","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [3,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","AI/ML","Pengumpulan Dataset","Mengumpulkan dataset hate speech & abusive language dari GitHub dan Kaggle","AI/ML Engineer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [4,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","AI/ML","EDA (Exploratory Data Analysis)","Menganalisis distribusi label; panjang teks; frekuensi kata; potensi bias dataset","AI/ML Engineer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [5,"Weekly #1","11/05/2026","14/05/2026","Planning & Setup","AI/ML","Relabeling Dataset (Awal)","Mengonversi label multi-kelas ke 2 kelas biner: PANTAS dan TIDAK PANTAS","AI/ML Engineer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [6,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","Backend","Setup FastAPI Project","Inisialisasi FastAPI; struktur folder; virtual environment; konfigurasi .env","Backend Developer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [7,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","Backend","Desain Database Schema","Merancang schema tabel: users; messages; rooms; violations (SQLAlchemy)","Backend Developer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [8,"Weekly #1","11/05/2026","17/05/2026","Planning & Setup","Frontend","Setup React Project","Inisialisasi React (Vite) + Tailwind CSS; konfigurasi folder components/pages/services","Frontend Developer","Belum Mulai","Tinggi","Repo & Environment Siap",""],
    [9,"Weekly #2","18/05/2026","24/05/2026","Development","AI/ML","Preprocessing Teks","Lowercasing; hapus URL/emoji; normalisasi slang; tokenisasi","AI/ML Engineer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [10,"Weekly #2","18/05/2026","24/05/2026","Development","AI/ML","Split Dataset","Membagi data train/validation/test (70/15/15) dengan stratified split","AI/ML Engineer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [11,"Weekly #2","18/05/2026","24/05/2026","Development","AI/ML","Relabeling Dataset (Finalisasi)","Finalisasi konversi label; pastikan semua bahasa kasar masuk TIDAK PANTAS","AI/ML Engineer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [12,"Weekly #2","18/05/2026","24/05/2026","Development","AI/ML","Fine-tuning IndoBERT (Iterasi 1)","Melatih indobert-base-p1 binary classification (PyTorch + HuggingFace)","AI/ML Engineer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [13,"Weekly #2","18/05/2026","24/05/2026","Development","Backend","Endpoint POST /api/messages","Menerima pesan dari FE; panggil AI inference; simpan ke DB; return respons","Backend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [14,"Weekly #2","18/05/2026","24/05/2026","Development","Backend","Endpoint GET /api/messages/{room_id}","Mengambil riwayat pesan sebuah room dengan pagination","Backend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [15,"Weekly #2","18/05/2026","24/05/2026","Development","Backend","WebSocket /ws/{room_id} (Awal)","Mulai membangun koneksi WebSocket real-time antar pengguna","Backend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [16,"Weekly #2","18/05/2026","24/05/2026","Development","Backend","Validasi & Keamanan","CORS setup; validasi input; sanitasi teks sebelum dikirim ke model","Backend Developer","Belum Mulai","Sedang","IndoBERT Iterasi 1 & API Moderasi",""],
    [17,"Weekly #2","18/05/2026","24/05/2026","Development","Frontend","Komponen ChatBubble","Pesan 2 varian: normal (PANTAS) atau tersembunyi (TIDAK PANTAS)","Frontend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [18,"Weekly #2","18/05/2026","24/05/2026","Development","Frontend","Komponen ChatInput","Input teks dengan tombol kirim; validasi input tidak kosong","Frontend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [19,"Weekly #2","18/05/2026","24/05/2026","Development","Frontend","Komponen ChatHeader","Header ruang chat (nama grup; jumlah anggota aktif)","Frontend Developer","Belum Mulai","Sedang","IndoBERT Iterasi 1 & API Moderasi",""],
    [20,"Weekly #2","18/05/2026","24/05/2026","Development","Frontend","Halaman Login","Form memilih nama pengguna simulasi sebelum masuk room chat","Frontend Developer","Belum Mulai","Sedang","IndoBERT Iterasi 1 & API Moderasi",""],
    [21,"Weekly #2","18/05/2026","21/05/2026","Development","Frontend","Halaman ChatRoom (Awal)","Gabungkan komponen chat; kelola state pesan; scroll otomatis","Frontend Developer","Belum Mulai","Tinggi","IndoBERT Iterasi 1 & API Moderasi",""],
    [22,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","AI/ML","Evaluasi Model","Hitung Accuracy/Precision/Recall/F1; Confusion Matrix; kalibrasi threshold","AI/ML Engineer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [23,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","AI/ML","Optimasi Model","Tuning hyperparameter; augmentasi data jika F1 < 82%","AI/ML Engineer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [24,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","AI/ML","Inference Function","Buat predict(text) -> {label, confidence} untuk dipanggil backend","AI/ML Engineer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [25,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","AI/ML","Streamlit Demo Endpoint","Demo interaktif: input teks -> label + confidence + visualisasi metrik","AI/ML Engineer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [26,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","AI/ML","Simpan & Dokumentasi Model","Upload model ke Google Drive; dokumentasikan di README","AI/ML Engineer","Belum Mulai","Sedang","Model Final & Streamlit Demo",""],
    [27,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Backend","Integrasi AI ke Backend","Muat model ke memori saat server start; panggil inference tiap pesan","Backend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [28,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Backend","Logika Aksi Moderasi","PANTAS -> normal; TIDAK PANTAS -> sembunyikan + notifikasi + log","Backend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [29,"Weekly #3","25/05/2026","28/05/2026","Development & Integration","Backend","WebSocket (Finalisasi)","Finalisasi dan uji WebSocket real-time stabil","Backend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [30,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Frontend","Halaman ChatRoom (Final)","Finalisasi integrasi komponen chat dengan WebSocket","Frontend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [31,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Frontend","Integrasi WebSocket (FE)","Hubungkan FE ke WebSocket BE untuk pesan real-time","Frontend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [32,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Frontend","Notifikasi Moderasi (FE)","Pesan edukatif untuk pengirim saat pesannya disembunyikan","Frontend Developer","Belum Mulai","Tinggi","Model Final & Streamlit Demo",""],
    [33,"Weekly #3","25/05/2026","31/05/2026","Development & Integration","Frontend","Halaman RoomList","Daftar room/grup tersedia untuk dipilih pengguna","Frontend Developer","Belum Mulai","Sedang","Model Final & Streamlit Demo",""],
    [34,"Weekly #4","01/06/2026","07/06/2026","Integration & Testing","Project Management","Koordinasi Integrasi & Testing","Pimpin integrasi FE-BE-AI; tulis test case; koordinasi bug fixing","Project Manager","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [35,"Weekly #4","01/06/2026","03/06/2026","Integration & Testing","Project Management","Laporan Kemajuan PIJAK","Submit laporan kemajuan ke PIJAK (DEADLINE: 3 Juni 2026)","Project Manager","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [36,"Weekly #4","01/06/2026","07/06/2026","Integration & Testing","Backend","Endpoint Admin","GET /api/admin/violations dan GET /api/admin/stats","Backend Developer","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [37,"Weekly #4","01/06/2026","07/06/2026","Integration & Testing","Backend","Testing & Dokumentasi API","Unit test; .env.example; dokumentasi API (auto-docs)","Backend Developer","Belum Mulai","Sedang","Integrasi Penuh & CP2 (3 Jun)",""],
    [38,"Weekly #4","01/06/2026","04/06/2026","Integration & Testing","Frontend","Halaman AdminDashboard","Log pelanggaran + grafik statistik (pie/bar chart)","Frontend Developer","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [39,"Weekly #4","01/06/2026","07/06/2026","Integration & Testing","Frontend","Integrasi Penuh FE-BE","Hubungkan semua komponen FE dengan endpoint API BE","Frontend Developer","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [40,"Weekly #4","01/06/2026","07/06/2026","Integration & Testing","All Team","Testing Integrasi FE-BE-AI","Uji alur end-to-end: kirim pesan -> moderasi -> respons -> tampilan","All Team","Belum Mulai","Tinggi","Integrasi Penuh & CP2 (3 Jun)",""],
    [41,"Weekly #5","08/06/2026","14/06/2026","Finalization","Project Management","README Documentation","Dokumentasi instalasi; cara menjalankan; link model; referensi dataset","Project Manager","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [42,"Weekly #5","08/06/2026","16/06/2026","Finalization","Project Management","Project Brief PIJAK","Susun Project Brief sesuai template PIJAK","Project Manager","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [43,"Weekly #5","14/06/2026","18/06/2026","Finalization","Project Management","Slide Presentasi","Slide PPTX/PDF: latar belakang; solusi; demo; hasil evaluasi","Project Manager","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [44,"Weekly #5","14/06/2026","18/06/2026","Finalization","Project Management","Video Presentasi","Rekam video 10 menit YouTube; semua anggota presentasi","Project Manager","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [45,"Weekly #5","08/06/2026","16/06/2026","Finalization","Project Management","Panduan Penggunaan","PDF atau video tutorial penggunaan aplikasi","Project Manager","Belum Mulai","Sedang","Submission Final (19 Jun)",""],
    [46,"Weekly #5","08/06/2026","14/06/2026","Finalization","All Team","Testing End-to-End","Pengujian menyeluruh termasuk edge case","All Team","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [47,"Weekly #5","12/06/2026","17/06/2026","Finalization","All Team","Bug Fixing","Perbaikan semua bug dari fase testing","All Team","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
    [48,"Weekly #5","08/06/2026","15/06/2026","Finalization","Frontend","Responsif Design","Pastikan tampilan baik di mobile (>=375px) dan desktop","Frontend Developer","Belum Mulai","Sedang","Submission Final (19 Jun)",""],
    [49,"Weekly #5","18/06/2026","19/06/2026","Submission","All Team","Final Submission","Submit semua deliverable (DEADLINE: 19 Juni 2026)","All Team","Belum Mulai","Tinggi","Submission Final (19 Jun)",""],
]

# Write headers
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

# Write data
for r_idx, row in enumerate(data, 2):
    cat = row[5]
    week = row[1]
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10)
        # Category color on col 6
        if c_idx == 6 and cat in CAT_COLORS:
            cell.fill = CAT_COLORS[cat]
    # Fill empty Catatan column
    if len(row) < len(headers):
        for c_idx in range(len(row)+1, len(headers)+1):
            cell = ws.cell(row=r_idx, column=c_idx, value="")
            cell.border = BORDER

# Data validation: Status dropdown
status_dv = DataValidation(type="list", formula1='"Belum Mulai,Sedang Dikerjakan,Selesai,Tertunda,Dibatalkan"', allow_blank=True)
status_dv.error = "Pilih status yang valid"
status_dv.errorTitle = "Status Tidak Valid"
ws.add_data_validation(status_dv)
status_dv.add(f"J2:J{len(data)+1}")

# Data validation: Prioritas dropdown
prio_dv = DataValidation(type="list", formula1='"Tinggi,Sedang,Rendah"', allow_blank=True)
ws.add_data_validation(prio_dv)
prio_dv.add(f"K2:K{len(data)+1}")

# Column widths
col_widths = [5, 12, 14, 14, 22, 20, 35, 55, 18, 18, 12, 35, 25]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top row
ws.freeze_panes = "A2"

# Auto filter
ws.auto_filter.ref = f"A1:M{len(data)+1}"

# Row height
ws.row_dimensions[1].height = 30
for r in range(2, len(data)+2):
    ws.row_dimensions[r].height = 35

# ============ SHEET 2: RINGKASAN MILESTONE ============
ws2 = wb.create_sheet("Milestone")
ms_headers = ["No","Minggu","Tanggal","Target Pencapaian","Deadline Penting","Progress (%)"]
ms_data = [
    [1,"Weekly #1","11-17 Mei 2026","Setup repo; dataset terkumpul; EDA selesai; skeleton FastAPI & React jalan","",0],
    [2,"Weekly #2","18-24 Mei 2026","IndoBERT iterasi pertama; endpoint API moderasi berfungsi; komponen chat UI selesai","",0],
    [3,"Weekly #3","25-31 Mei 2026","Model dievaluasi & dioptimasi; WebSocket berfungsi; Streamlit demo online; integrasi AI-BE","",0],
    [4,"Weekly #4","1-7 Juni 2026","Integrasi penuh FE-BE-AI; dashboard admin berfungsi; laporan kemajuan PIJAK","3 Juni - Laporan PIJAK",0],
    [5,"Weekly #5","8-19 Juni 2026","Testing E2E; bug fixing; README; Project Brief; slide; video presentasi","19 Juni - DEADLINE FINAL",0],
]

for col, h in enumerate(ms_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

for r_idx, row in enumerate(ms_data, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=11)

ms_widths = [5, 12, 20, 60, 30, 15]
for i, w in enumerate(ms_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 30
for r in range(2, 7):
    ws2.row_dimensions[r].height = 50

# Progress dropdown (0-100 in steps of 10)
prog_dv = DataValidation(type="list", formula1='"0,10,20,30,40,50,60,70,80,90,100"', allow_blank=True)
ws2.add_data_validation(prog_dv)
prog_dv.add("F2:F6")

# ============ SHEET 3: GANTT CHART ============
ws3 = wb.create_sheet("Gantt Chart")

# Project dates: May 11 - June 19
start_date = datetime(2026, 5, 11)
end_date = datetime(2026, 6, 19)
total_days = (end_date - start_date).days + 1

# Headers
ws3.cell(row=1, column=1, value="Task").font = HEADER_FONT
ws3.cell(row=1, column=1).fill = HEADER_FILL
ws3.cell(row=1, column=1).border = BORDER
ws3.cell(row=1, column=2, value="PIC").font = HEADER_FONT
ws3.cell(row=1, column=2).fill = HEADER_FILL
ws3.cell(row=1, column=2).border = BORDER

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 18

# Date headers
for d in range(total_days):
    col = d + 3
    dt = start_date + timedelta(days=d)
    cell = ws3.cell(row=1, column=col, value=dt.strftime("%d/%m"))
    cell.font = Font(name="Calibri", size=7, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
    cell.border = BORDER
    ws3.column_dimensions[get_column_letter(col)].width = 3.5

    # Weekend shading
    if dt.weekday() >= 5:
        for future_row in range(2, len(data)+2):
            c = ws3.cell(row=future_row, column=col)
            c.fill = PatternFill("solid", fgColor="F3F4F6")

GANTT_COLORS = {
    "Project Management": PatternFill("solid", fgColor="3B82F6"),
    "AI/ML": PatternFill("solid", fgColor="F59E0B"),
    "Backend": PatternFill("solid", fgColor="10B981"),
    "Frontend": PatternFill("solid", fgColor="8B5CF6"),
    "All Team": PatternFill("solid", fgColor="EC4899"),
}

for r_idx, row in enumerate(data, 2):
    task_name = row[6]
    pic = row[8]
    cat = row[5]
    t_start = datetime.strptime(row[2], "%d/%m/%Y")
    t_end = datetime.strptime(row[3], "%d/%m/%Y")

    ws3.cell(row=r_idx, column=1, value=task_name).font = Font(name="Calibri", size=9)
    ws3.cell(row=r_idx, column=1).border = BORDER
    ws3.cell(row=r_idx, column=1).alignment = Alignment(vertical="center")
    ws3.cell(row=r_idx, column=2, value=pic).font = Font(name="Calibri", size=8)
    ws3.cell(row=r_idx, column=2).border = BORDER

    for d in range(total_days):
        col = d + 3
        dt = start_date + timedelta(days=d)
        cell = ws3.cell(row=r_idx, column=col)
        cell.border = Border(
            left=Side(style="hair", color="E5E7EB"),
            right=Side(style="hair", color="E5E7EB"),
            top=Side(style="hair", color="E5E7EB"),
            bottom=Side(style="hair", color="E5E7EB")
        )
        if t_start <= dt <= t_end:
            cell.fill = GANTT_COLORS.get(cat, PatternFill("solid", fgColor="9CA3AF"))

    ws3.row_dimensions[r_idx].height = 20

ws3.freeze_panes = "C2"
ws3.row_dimensions[1].height = 50

# ============ SAVE ============
filepath = r"C:\Users\Ahsani Fadhli Ilahi\Downloads\PROYEK-AI-PIJAK-AHSANI-TEAM\Timeline-AI-SHIELD.xlsx"
wb.save(filepath)
print(f"Excel file saved: {filepath}")
print("Features: 3 sheets (Timeline, Milestone, Gantt Chart), dropdown menus, auto-filter, color coding, freeze panes")
